import pandas as pd
import os
import re
import matplotlib.pyplot as plt
import seaborn as sns
from glob import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_xy_from_filename(filename):
    """
    ดึงค่าพิกัด x y จากชื่อไฟล์ที่อยู่ในรูปแบบ ..._x y... เช่น _-35 -69
    """
    match = re.search(r'_(-?\d+)\s(-?\d+)', filename)
    if match:
        x = int(match.group(1))
        y = int(match.group(2))
        return x, y
    return None, None


folder_path = 'Y:\XXXXX' #folder ต้นทาง
csv_files = glob(os.path.join(folder_path, '*.csv'))

combined_data = []

for file in csv_files:
    filename = os.path.basename(file)
    x, y = extract_xy_from_filename(filename)

    if x is None or y is None:
        continue  

    try:
        
        with open(file, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
            if len(lines) >= 256:
                line256 = lines[255].strip()
                parts = re.split(r'[,\t\s]+', line256)  # แยกด้วย comma, tab หรือ space
                if len(parts) > 2:
                    value = float(parts[2])
    except Exception as e:
        print(f"Error in {filename}: {e}")
        value = None

    combined_data.append({
        'filename': filename,
        'x': x,
        'y': y,
        'value': value
    })


df_combined = pd.DataFrame(combined_data).sort_values(by=['x', 'y'])

vmin = df_combined['value'].min()
vmax = df_combined['value'].max()
print(f"Min value: {vmin}")
print(f"Max value: {vmax}")

df_combined['abs_x'] = df_combined['x'].abs() #convert x y value to abs
df_combined['abs_y'] = df_combined['y'].abs()

folder_name = os.path.basename(folder_path)
output_excel = os.path.join(r'Z:\XXXXX', f"{folder_name}.xlsx")
df_combined.to_excel(output_excel, index=False) #convert dataframe to excel output file
print(f'success: {output_excel}')
#print("data:", combined_data[:3]) 

pivot = df_combined.pivot(index='abs_y', columns='abs_x', values='value')

# ฟังก์ชันแปลงค่าตัวเลขเป็นสี (gradient สีฟ้า → เขียว → เหลือง → แดง)
def value_to_color(val):
    norm = (val - vmin) / (vmax - vmin + 1e-9)
    r = int(255 * norm)
    g = int(255 * (1 - norm))
    b = 0
    return f"{r:02X}{g:02X}{b:02X}"

# สร้าง workbook
wb = Workbook()
ws = wb.active
ws.title = "Heatmap"

# เขียนชื่อ column
ws.append([""] + list(pivot.columns))

# เขียนค่าพร้อมสีแต่ละแถว
for abs_y, row in pivot.iterrows():
    row_data = [abs_y]
    for abs_x in pivot.columns:
        val = row[abs_x]
        row_data.append(val)
    ws.append(row_data)

# ใส่สีใน cell
for i, abs_y in enumerate(pivot.index, start=2):
    for j, abs_x in enumerate(pivot.columns, start=2):
        val = pivot.loc[abs_y, abs_x]
        if pd.notnull(val):
            color = value_to_color(val)
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=i, column=j).fill = fill
            
# Save ไฟล์
wb.save(rf"Z:\data\{folder_name}_heatmap.xlsx")
print(" heatmap to Excel ")

plt.figure(figsize=(10, 8))
vmin = df_combined['value'].quantile(0.05)  # 5th percentile
vmax = df_combined['value'].quantile(0.95)  # 95th percentile
print(f"VMin: {vmin}, VMax: {vmax}")
sns.heatmap(pivot, cmap='coolwarm', annot=True,annot_kws={"size": 6}, fmt=".2f", linewidths=0.5, vmin=vmin, vmax=vmax )

plt.title(f"Heatmap of {folder_name} at (x, y) Coordinates")
plt.xlabel("x")
plt.ylabel("y")
plt.tight_layout()
plt.show()